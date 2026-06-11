from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from typing import List, Dict


def export_knowledge_to_excel(data: List[Dict]) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "知识清单"

    headers = ["ID", "标题", "分类", "提交人", "状态", "创建时间", "更新时间"]
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="165DFF", end_color="165DFF", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    status_map = {
        "pending": "待复核",
        "approved": "已通过",
        "rejected": "已驳回"
    }

    for row, item in enumerate(data, 2):
        ws.cell(row=row, column=1, value=item.get("id", ""))
        ws.cell(row=row, column=2, value=item.get("title", ""))
        ws.cell(row=row, column=3, value=item.get("category_name", ""))
        ws.cell(row=row, column=4, value=item.get("submitter_name", ""))
        ws.cell(row=row, column=5, value=status_map.get(item.get("review_status", ""), item.get("review_status", "")))
        ws.cell(row=row, column=6, value=str(item.get("created_at", "")))
        ws.cell(row=row, column=7, value=str(item.get("updated_at", "")))

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[chr(64 + col)].width = 20

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
